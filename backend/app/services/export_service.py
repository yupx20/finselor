"""Export service — generate CSV and XLSX files from transaction data."""

import csv
import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_csv(transactions: list[dict]) -> io.StringIO:
    """Generate a CSV file from transaction data."""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    # Header
    writer.writerow(["Date", "Type", "Category", "Amount", "Notes"])

    for t in transactions:
        writer.writerow([
            t.get("trx_date", ""),
            t.get("trx_type", ""),
            t.get("category_name", ""),
            t.get("amount", 0),
            t.get("notes", "") or "",
        ])

    output.seek(0)
    return output


def generate_xlsx(transactions: list[dict], month: int, year: int) -> io.BytesIO:
    """Generate a styled XLSX file from transaction data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Transactions {year}-{month:02d}"

    # Styles
    header_font = Font(name="Inter", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    income_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    expense_fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")

    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # Title row
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Finselor — Transaction Report ({year}-{month:02d})"
    title_cell.font = Font(name="Inter", bold=True, size=14, color="0F172A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row
    headers = ["Date", "Type", "Category", "Amount (IDR)", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    total_income = Decimal("0")
    total_expense = Decimal("0")

    for row_idx, t in enumerate(transactions, 4):
        trx_type = t.get("trx_type", "")
        amount = Decimal(str(t.get("amount", 0)))

        if trx_type == "INCOME":
            total_income += amount
            row_fill = income_fill
        else:
            total_expense += amount
            row_fill = expense_fill

        row_data = [
            t.get("trx_date", ""),
            trx_type,
            t.get("category_name", ""),
            float(amount),
            t.get("notes", ""),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = border
            if col_idx == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # Summary row
    summary_row = len(transactions) + 5
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=11)
    ws.cell(row=summary_row + 1, column=1, value="Total Income:")
    ws.cell(row=summary_row + 1, column=2, value=float(total_income)).number_format = '#,##0.00'
    ws.cell(row=summary_row + 2, column=1, value="Total Expense:")
    ws.cell(row=summary_row + 2, column=2, value=float(total_expense)).number_format = '#,##0.00'
    ws.cell(row=summary_row + 3, column=1, value="Surplus:").font = Font(bold=True)
    ws.cell(row=summary_row + 3, column=2, value=float(total_income - total_expense)).number_format = '#,##0.00'
    ws.cell(row=summary_row + 3, column=2).font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
